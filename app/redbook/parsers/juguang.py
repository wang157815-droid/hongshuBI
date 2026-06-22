import openpyxl

from app.redbook.constants.source_type import SOURCE_JUGUANG
from app.schemas.redbook import ParseResult

from .base import BaseRedbookParser, clean_text, compact_raw, to_date, to_decimal, to_int


class JuguangParser(BaseRedbookParser):
    source_type = SOURCE_JUGUANG

    FIELD_MAP = {
        "时间": ("stat_date", to_date),
        "创意名称": ("creative_name", clean_text),
        "创意ID": ("creative_id", clean_text),
        "笔记/素材ID": ("note_id", clean_text),
        "笔记/素材链接": ("note_url", clean_text),
        "单元名称": ("unit_name", clean_text),
        "单元ID": ("unit_id", clean_text),
        "计划名称": ("plan_name", clean_text),
        "计划ID": ("plan_id", clean_text),
        "创作者": ("creator_name", clean_text),
        "消费": ("cost", to_decimal),
        "展现量": ("impressions", to_int),
        "点击量": ("clicks", to_int),
        "点击率": ("ctr", lambda v: to_decimal(v, is_percent=True)),
        "平均点击成本": ("cpc", to_decimal),
        "平均千次展现费用": ("cpm", to_decimal),
        "点赞": ("likes", to_int),
        "收藏": ("collects", to_int),
        "评论": ("comments", to_int),
        "关注": ("follows", to_int),
        "分享": ("shares", to_int),
        "互动量": ("interactions", to_int),
        "平均互动成本": ("cpe", to_decimal),
        "行动按钮点击量": ("action_button_clicks", to_int),
        "行动按钮点击率": ("action_button_ctr", lambda v: to_decimal(v, is_percent=True)),
        "截图": ("screenshots", to_int),
        "保存图片": ("save_images", to_int),
        "搜索组件点击量": ("search_component_clicks", to_int),
        "搜索组件点击转化率": ("search_component_ctr", lambda v: to_decimal(v, is_percent=True)),
        "搜后阅读量": ("post_search_reads", to_int),
        "小红星站外活跃行为UV(30日归因)": ("offsite_active_uv_30d", to_int),
        "小红星站外活跃成本(30日归因)": ("offsite_active_cost_30d", to_decimal),
        "新增种草人群": ("new_seed_users", to_int),
        "新增种草人群成本": ("new_seed_user_cost", to_decimal),
        "新增深度种草人群": ("new_deep_seed_users", to_int),
        "新增深度种草人群成本": ("new_deep_seed_user_cost", to_decimal),
    }
    REQUIRED_HEADERS = {"时间", "笔记/素材ID", "消费", "展现量", "点击量", "互动量"}

    def parse(self) -> ParseResult:
        workbook = openpyxl.load_workbook(self.file_path, data_only=True, read_only=False)
        sheet = workbook["创意-投放数据"] if "创意-投放数据" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_index = {header: idx for idx, header in enumerate(headers) if header}
        missing = sorted(self.REQUIRED_HEADERS - set(header_index))
        if missing:
            for header in missing:
                self.error(1, header, None, "missing_header", f"缺少必填字段：{header}")
            return ParseResult(source_type=self.source_type, total_rows=sheet.max_row, errors=self.errors)

        records = []
        summary = {}
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            first_cell = clean_text(row[0] if row else None)
            if not first_cell:
                continue
            raw = {header: row[idx] for header, idx in header_index.items() if idx < len(row)}
            if first_cell.startswith("合计"):
                summary = compact_raw(raw)
                continue
            record = {"row_number": row_number}
            for source_header, (target_field, converter) in self.FIELD_MAP.items():
                idx = header_index.get(source_header)
                if idx is not None and idx < len(row):
                    record[target_field] = converter(row[idx])
            if not record.get("note_id"):
                self.warn(row_number, "笔记/素材ID", None, "missing_note_id", "笔记/素材ID为空")
            record["raw_json"] = compact_raw(raw)
            records.append(record)

        total_rows = sheet.max_row
        workbook.close()
        return ParseResult(
            source_type=self.source_type,
            total_rows=total_rows,
            data_rows=len(records),
            success_rows=len(records),
            failed_rows=len(self.errors),
            warnings=self.warnings,
            errors=self.errors,
            records=records,
            summary=summary,
        )
