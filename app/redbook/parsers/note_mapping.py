import openpyxl

from app.redbook.constants.source_type import SOURCE_NOTE_MAPPING
from app.schemas.redbook import ParseResult

from .base import BaseRedbookParser, clean_text, compact_raw, to_date


class NoteMappingParser(BaseRedbookParser):
    source_type = SOURCE_NOTE_MAPPING

    FIELD_MAP = {
        "笔记ID": ("note_id", clean_text),
        "达人": ("blogger_name", clean_text),
        "笔记类型": ("note_type", clean_text),
        "达人分类": ("blogger_type", clean_text),
        "笔记短链接": ("note_url", clean_text),
        "发布时间": ("publish_date", to_date),
        "内容方向": ("content_direction", clean_text),
        "所属产品": ("product_category", clean_text),
    }
    REQUIRED_HEADERS = {"笔记ID", "达人", "笔记类型", "达人分类", "内容方向", "所属产品"}

    def parse(self) -> ParseResult:
        workbook = openpyxl.load_workbook(self.file_path, data_only=True, read_only=False)
        sheet = workbook["匹配源"] if "匹配源" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_index = {header: idx for idx, header in enumerate(headers) if header}
        missing = sorted(self.REQUIRED_HEADERS - set(header_index))
        if missing:
            for header in missing:
                self.error(1, header, None, "missing_header", f"缺少必填字段：{header}")
            return ParseResult(source_type=self.source_type, total_rows=sheet.max_row, errors=self.errors)

        records = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw = {header: row[idx] for header, idx in header_index.items() if idx < len(row)}
            if not any(value not in (None, "") for value in raw.values()):
                continue
            record = {"row_number": row_number, "status": "active"}
            for source_header, (target_field, converter) in self.FIELD_MAP.items():
                idx = header_index.get(source_header)
                if idx is not None and idx < len(row):
                    record[target_field] = converter(row[idx])
            record["product_name"] = record.get("product_category")
            if not record.get("note_id"):
                self.warn(row_number, "笔记ID", None, "missing_note_id", "笔记ID为空")
                continue
            record["raw_json"] = compact_raw(raw)
            records.append(record)

        total_rows = sheet.max_row
        sheet_name = sheet.title
        workbook.close()
        return ParseResult(
            source_type=self.source_type,
            total_rows=total_rows,
            data_rows=len(records),
            success_rows=len(records),
            failed_rows=len(self.errors),
            warnings=self.warnings,
            errors=self.errors,
            records=records,
            summary={"sheet_name": sheet_name},
        )
