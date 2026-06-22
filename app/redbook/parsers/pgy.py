import openpyxl

from app.redbook.constants.source_type import SOURCE_PGY
from app.schemas.redbook import ParseResult

from .base import BaseRedbookParser, clean_text, compact_raw, to_date, to_decimal, to_int


def first_valid_header_value(headers, row, header, converter=clean_text):
    for idx, item in enumerate(headers):
        if item != header or idx >= len(row):
            continue
        value = converter(row[idx])
        if value not in (None, ""):
            return value
    return None


class PgyParser(BaseRedbookParser):
    source_type = SOURCE_PGY

    FIELD_MAP = {
        "数据更新日期": ("update_date", to_date),
        "博主昵称": ("blogger_name", clean_text),
        "博主主页链接": ("blogger_home_url", clean_text),
        "博主粉丝量": ("fans_count", to_int),
        "博主健康等级": ("health_level", clean_text),
        "笔记标题": ("note_title", clean_text),
        "笔记链接": ("note_url", clean_text),
        "笔记类型": ("note_type", clean_text),
        "笔记发布日期": ("publish_date", to_date),
        "笔记来源": ("note_source", clean_text),
        "笔记id": ("note_id", clean_text),
        "内容标签": ("content_tag", clean_text),
        "订单id": ("order_id", clean_text),
        "合作名称": ("cooperation_name", clean_text),
        "报备品牌": ("brand_name", clean_text),
        "下单账号": ("account_name", clean_text),
        "博主报价": ("blogger_quote_amount", to_decimal),
        "服务费金额": ("service_fee_amount", to_decimal),
        "是否为优效模式": ("is_effective_mode", clean_text),
        "spu名称": ("spu_name", clean_text),
        "曝光量": ("exposure", to_int),
        "阅读量": ("read_count", to_int),
        "阅读UV": ("read_uv", to_int),
        "互动量": ("interaction_count", to_int),
        "互动率": ("interaction_rate", lambda v: to_decimal(v, is_percent=True)),
        "点赞量": ("like_count", to_int),
        "收藏量": ("collect_count", to_int),
        "评论量": ("comment_count", to_int),
        "分享量": ("share_count", to_int),
        "关注量": ("follow_count", to_int),
        "自然曝光量": ("natural_exposure", to_int),
        "自然阅读量": ("natural_read_count", to_int),
        "推广曝光量": ("promotion_exposure", to_int),
        "推广阅读量": ("promotion_read_count", to_int),
        "加热曝光量": ("heat_exposure", to_int),
        "加热阅读量": ("heat_read_count", to_int),
    }
    REQUIRED_HEADERS = {"博主昵称", "笔记标题", "笔记id", "博主报价", "服务费金额"}

    def parse(self) -> ParseResult:
        workbook = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)
        sheet = workbook["笔记批量数据"] if "笔记批量数据" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        first_level = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        second_level = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=3, max_row=3))]
        headers = []
        for idx, second in enumerate(second_level):
            headers.append(second or (first_level[idx] if idx < len(first_level) else None))
        header_index = {header: idx for idx, header in enumerate(headers) if header}
        missing = sorted(self.REQUIRED_HEADERS - set(header_index))
        if missing:
            for header in missing:
                self.error(3, header, None, "missing_header", f"缺少必填字段：{header}")
            return ParseResult(source_type=self.source_type, total_rows=sheet.max_row, errors=self.errors)

        records = []
        wide_records = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            raw = {header: row[idx] for idx, header in enumerate(headers) if header and idx < len(row)}
            if not any(value not in (None, "") for value in raw.values()):
                continue
            standard = {"row_number": row_number}
            for source_header, (target_field, converter) in self.FIELD_MAP.items():
                idx = header_index.get(source_header)
                if idx is not None and idx < len(row):
                    standard[target_field] = converter(row[idx])
            standard["task_id"] = first_valid_header_value(headers, row, "任务ID")
            if not standard.get("note_id"):
                self.warn(row_number, "笔记id", None, "missing_note_id", "笔记id为空")
            raw_json = compact_raw(raw)
            standard["raw_json"] = raw_json
            records.append(standard)
            wide_records.append(
                {
                    "row_number": row_number,
                    "update_date": standard.get("update_date"),
                    "note_id": standard.get("note_id"),
                    "raw_json": raw_json,
                }
            )

        total_rows = sheet.max_row
        sheet_name = sheet.title
        workbook.close()
        return ParseResult(
            source_type=self.source_type,
            total_rows=total_rows,
            data_rows=len(records),
            success_rows=len(records),
            failed_rows=len(self.errors),
            warnings=self.warnings,
            errors=self.errors,
            records=records,
            wide_records=wide_records,
            summary={"sheet_name": sheet_name},
        )
