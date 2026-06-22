from decimal import Decimal
from zipfile import ZipFile
import re
import xml.etree.ElementTree as ET

import openpyxl

from app.redbook.constants.source_type import SOURCE_KEYWORD_SEARCH
from app.schemas.redbook import ParseResult

from .base import BaseRedbookParser, clean_text, compact_raw, to_date, to_decimal


class KeywordSearchParser(BaseRedbookParser):
    source_type = SOURCE_KEYWORD_SEARCH
    parser_version = "v1"

    def parse(self) -> ParseResult:
        try:
            rows = self._read_with_openpyxl()
            reader = "openpyxl"
        except Exception as exc:
            self.warn(None, None, None, "openpyxl_fallback", f"openpyxl读取失败，已尝试XML兜底：{exc}")
            rows = self._read_with_xml()
            reader = "xml"

        if not rows:
            self.error(1, None, None, "empty_sheet", "搜索日报为空")
            return ParseResult(source_type=self.source_type, errors=self.errors)

        headers = rows[0]
        date_columns = []
        for index, value in enumerate(headers[1:], start=2):
            stat_date = to_date(value)
            if stat_date:
                date_columns.append((index, stat_date))
        if not date_columns:
            self.error(1, None, None, "missing_date_columns", "搜索日报未识别到日期列")
            return ParseResult(source_type=self.source_type, total_rows=len(rows), errors=self.errors)

        records = []
        keyword_count = 0
        less_than_count = 0
        for row_number, row in enumerate(rows[1:], start=2):
            keyword = clean_text(row[0] if row else None)
            if not keyword:
                continue
            keyword_count += 1
            product_category, keyword_type = classify_keyword(keyword)
            for column_number, stat_date in date_columns:
                raw_cell = row[column_number - 1] if column_number - 1 < len(row) else None
                raw_value = clean_text(raw_cell)
                if raw_value is None:
                    continue
                index_value, is_less_than, threshold_value = parse_search_value(raw_value)
                if is_less_than:
                    less_than_count += 1
                records.append(
                    {
                        "row_number": row_number,
                        "column_number": column_number,
                        "stat_date": stat_date,
                        "keyword": keyword,
                        "product_category": product_category,
                        "keyword_type": keyword_type,
                        "raw_value": raw_value,
                        "search_index": index_value,
                        "is_less_than_threshold": is_less_than,
                        "threshold_value": threshold_value,
                        "estimate_value": None,
                        "raw_json": compact_raw(
                            {
                                "keyword": keyword,
                                "date_header": headers[column_number - 1],
                                "raw_value": raw_value,
                            }
                        ),
                    }
                )

        return ParseResult(
            source_type=self.source_type,
            total_rows=len(rows),
            data_rows=len(records),
            success_rows=len(records),
            failed_rows=len(self.errors),
            warnings=self.warnings,
            errors=self.errors,
            records=records,
            summary={
                "reader": reader,
                "keyword_count": keyword_count,
                "date_count": len(date_columns),
                "less_than_threshold_count": less_than_count,
            },
        )

    def _read_with_openpyxl(self):
        workbook = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
        workbook.close()
        return rows

    def _read_with_xml(self):
        with ZipFile(self.file_path) as archive:
            shared_strings = read_shared_strings(archive)
            sheet_names = read_sheet_names(archive)
            first_sheet_path = sheet_names[0][1] if sheet_names else "xl/worksheets/sheet1.xml"
            if not first_sheet_path.startswith("xl/"):
                first_sheet_path = f"xl/{first_sheet_path}"
            xml = archive.read(first_sheet_path)
        root = ET.fromstring(xml)
        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        parsed_rows = []
        for row_node in root.findall(".//main:sheetData/main:row", ns):
            row_index = int(row_node.attrib.get("r", len(parsed_rows) + 1))
            while len(parsed_rows) < row_index:
                parsed_rows.append([])
            row_values = parsed_rows[row_index - 1]
            for cell in row_node.findall("main:c", ns):
                ref = cell.attrib.get("r", "")
                col_index = column_index_from_ref(ref)
                while len(row_values) < col_index:
                    row_values.append(None)
                row_values[col_index - 1] = cell_value(cell, shared_strings, ns)
        return parsed_rows


def parse_search_value(value):
    text = clean_text(value)
    if text is None:
        return None, False, None
    less_than_match = re.match(r"^<\s*(\d+(?:\.\d+)?)$", text)
    if less_than_match:
        return None, True, Decimal(less_than_match.group(1))
    return to_decimal(text), False, None


def classify_keyword(keyword: str):
    lower_keyword = keyword.lower()
    if "缦拉" in keyword or "mangla" in lower_keyword:
        return category_for_keyword(keyword), "brand"
    competitors = (
        "观夏",
        "闻献",
        "呈白",
        "边顶",
        "龟宝香居",
        "阿蒂仙",
        "emonster",
        "aesop",
        "diptyque",
        "byredo",
        "melt season",
        "cottee",
        "寓义",
        "le labo",
        "momaek",
    )
    if any(item in lower_keyword or item in keyword for item in competitors):
        return category_for_keyword(keyword), "competitor"
    category = category_for_keyword(keyword)
    if category:
        return category, "category" if keyword == category else "product"
    return None, "custom"


def category_for_keyword(keyword: str):
    if "香水" in keyword:
        return "香水"
    if "线香" in keyword or "藏香" in keyword or keyword == "线香":
        return "线香"
    return None


def read_shared_strings(archive: ZipFile):
    try:
        xml = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ET.fromstring(xml)
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    values = []
    for item in root.findall("main:si", ns):
        parts = [node.text or "" for node in item.findall(".//main:t", ns)]
        values.append("".join(parts))
    return values


def read_sheet_names(archive: ZipFile):
    try:
        workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    except KeyError:
        return []
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    rels = {item.attrib["Id"]: item.attrib["Target"] for item in rels_xml.findall("pkg:Relationship", ns)}
    sheets = []
    for sheet in workbook_xml.findall("main:sheets/main:sheet", ns):
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = rels.get(rel_id)
        if target:
            sheets.append((sheet.attrib.get("name"), target))
    return sheets


def cell_value(cell, shared_strings, ns):
    cell_type = cell.attrib.get("t")
    value_node = cell.find("main:v", ns)
    if value_node is None:
        return None
    value = value_node.text
    if cell_type == "s":
        index = int(value)
        return shared_strings[index] if index < len(shared_strings) else None
    if cell_type == "b":
        return value == "1"
    if value and re.match(r"^\d+(\.\d+)?$", value):
        number = Decimal(value)
        if number == number.to_integral():
            return int(number)
        return number
    return value


def column_index_from_ref(ref: str):
    letters = re.match(r"([A-Z]+)", ref).group(1)
    index = 0
    for letter in letters:
        index = index * 26 + ord(letter) - ord("A") + 1
    return index
